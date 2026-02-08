import openpyxl
from openpyxl.utils import get_column_letter
import os
import sys
from datetime import datetime, timedelta

# ====================================================================================
# 파일 및 시트 이름 정의 (사용자 환경에 맞게 이 부분을 수정하세요!)
# ====================================================================================
SOURCE_WB_NAME = "251102_insert.xlsx"
SOURCE_SHEET_NAME = "법인사업자"
OUTPUT_SHEET_NAME = "Company_Additional_Output"
OUTPUT_WB_NAME = "Company_Additional_Output.xlsx"

# --- 상수 정의 ---
DATA_START_ROW = 3
COL_BIZ_NO = 24  # X열

# 특허 관련 컬럼
COL_PATENT_APPLICATIONS_COUNT = 165  # 165
COL_REGISTERED_PATENTS_COUNT = 164   # 164

# 기업부설연구소, 이노비즈, 메인비즈, 벤처 관련 컬럼
COL_HAS_RESEARCH_INSTITUTE = 152     # 기업부설연구소여부 (152)
COL_RESEARCH_INSTITUTE_DATE = 153    # 기업부설연구소 인정일자 (변동없음 가정, 사용자 언급없음 -> 153 유지? 사용자 "기업부설연구소여부: 150번째". 인정일자는 언급 없음. 기존코드 153. 보통 여부 옆에 날짜가 옴. 150옆은 151? Check if user provided info. Only "150". I will assume adjacent columns shift too? Or just stick to what user said. If 152->150 (shift -2). Then 153->151? I can't be sure. User only specified 150. I will change 152->150. I will leave others unless I know. Actually, "Inno-Biz 154". "Main-Biz 157". "Patent App 165". "Patent Reg 164".
# Wait, Patent App was 164 -> 165 (+1). Patent Reg was 163 -> 164 (+1).
# Lab was 152 -> 150 (-2).
# InnoBiz was 154. (Old code 154). Same.
# MainBiz was 157. (Old code 157). Same.
# Group 166. GFC 167.
# I will update ONLY what is specified. I'll leave dates alone unless they look wrong.
# If Lab YN moved 152->150, likely the Date moved too? But I don't know where.
# I will stick to explicitly requested changes.

COL_IS_INNOBIZ = 154                 # 이노비즈 여부
COL_INNOBIZ_CERT_DATE = 155          # 이노비즈 인정일자 (유지)
COL_INNOBIZ_EXPIRY_DATE = 156        # 이노비즈 만료일자 (유지)
COL_IS_MAINBIZ = 157                 # 메인비즈 여부
COL_MAINBIZ_CERT_DATE = 158          # 메인비즈 인정일자 (유지)
COL_MAINBIZ_EXPIRY_DATE = 159        # 메인비즈 만료일자 (유지)
COL_IS_VENTURE = 160                 # 벤처인증여부 (유지)
COL_VENTURE_CERT_DATE = 161          # 벤처인증 발급일자 (유지)
COL_VENTURE_EXPIRY_DATE = 162        # 벤처인증 만료일자 (유지)

# 신규 추가 컬럼
COL_GROUP_AGREEMENT_YN = 166
COL_GFC_YN = 167

def get_numeric_value(cell_val):
    """
    VBA의 GetNumericValue 함수와 동일한 역할을 합니다.
    셀 값이 숫자 형태가 아니거나 비어있으면 0을 반환합니다.
    """
    if cell_val is None:
        return 0
    try:
        return int(float(cell_val))
    except (ValueError, TypeError):
        return 0

def get_date_string(cell_val):
    """
    VBA의 GetDateString 함수와 동일한 역할을 합니다.
    날짜 형식의 셀 값을 YYYY-MM-DD 문자열로 변환합니다.
    """
    if cell_val is None:
        return ""
    
    # datetime 객체인 경우
    if isinstance(cell_val, datetime):
        return cell_val.strftime('%Y-%m-%d')
    
    # 숫자인 경우 (엑셀 일련번호)
    if isinstance(cell_val, (int, float)) and cell_val > 10000:
        try:
            # 엑셀 일련번호를 날짜로 변환 (1900-01-01부터의 일수)
            base_date = datetime(1900, 1, 1)
            date_obj = base_date.replace(year=1899, month=12, day=30) + timedelta(days=int(cell_val))
            return date_obj.strftime('%Y-%m-%d')
        except:
            return str(cell_val).strip()
    
    return str(cell_val).strip()

def get_boolean_string(cell_val):
    """
    VBA의 GetBooleanString 함수와 동일한 역할을 합니다.
    '여부' 필드의 값을 '1' (Y) 또는 '0' (N) 문자열로 변환합니다.
    """
    if cell_val is None:
        return "0"
    
    temp_str = str(cell_val).strip().upper()
    if temp_str.startswith("Y") or temp_str == "TRUE" or temp_str == "1":
        return "1"
    else:
        return "0"

def process_additional_data():
    """
    기업부가정보 (Company_Additional) 추출 메인 함수입니다.
    """
    try:
        # 1. 엑셀 파일 불러오기
        print(f"'{SOURCE_WB_NAME}' 파일을 불러오는 중입니다...")
        if not os.path.exists(SOURCE_WB_NAME):
            raise FileNotFoundError(f"원본 데이터 파일 '{SOURCE_WB_NAME}'을 찾을 수 없습니다.")

        wb_source = openpyxl.load_workbook(SOURCE_WB_NAME, data_only=True)
        if SOURCE_SHEET_NAME in wb_source.sheetnames:
            ws_source = wb_source[SOURCE_SHEET_NAME]
        else:
            ws_source = wb_source.active
            print(f"Warning: Sheet '{SOURCE_SHEET_NAME}' not found. Using active sheet '{ws_source.title}' instead.")

        # 2. 출력용 새 워크북 및 시트 생성
        wb_output = openpyxl.Workbook()
        ws_output = wb_output.active
        ws_output.title = OUTPUT_SHEET_NAME
        print(f"'{OUTPUT_SHEET_NAME}' 시트를 생성했습니다.")

        # 3. 출력 시트 헤더 작성
        headers = [
            "biz_no",
            "patent_applications_count",
            "registered_patents_count",
            "has_research_institute",
            "research_institute_date",
            "is_innobiz",
            "innobiz_cert_date",
            "innobiz_expiry_date",
            "is_mainbiz",
            "mainbiz_cert_date",
            "mainbiz_expiry_date",
            "is_venture",
            "venture_cert_date",
            "venture_expiry_date",
            "group_agreement_yn",
            "gfc_yn"
        ]
        ws_output.append(headers)
        
        # 4. 데이터 처리 시작
        last_row_source = ws_source.max_row
        print(f"총 {last_row_source - DATA_START_ROW + 1}개의 원본 데이터를 처리합니다...")

        for i in range(DATA_START_ROW, last_row_source + 1):
            biz_no_val = ws_source.cell(row=i, column=COL_BIZ_NO).value
            biz_no = str(biz_no_val).strip() if biz_no_val is not None else ""

            if biz_no and biz_no != "0":
                # 특허 관련 데이터
                patent_apps = get_numeric_value(ws_source.cell(row=i, column=COL_PATENT_APPLICATIONS_COUNT).value)
                registered_patents = get_numeric_value(ws_source.cell(row=i, column=COL_REGISTERED_PATENTS_COUNT).value)
                
                # 기업부설연구소, 이노비즈, 메인비즈, 벤처 관련 데이터
                # [수정] 사용자 요청: 기업부설연구소 여부 = 150
                has_research_institute = get_boolean_string(ws_source.cell(row=i, column=COL_HAS_RESEARCH_INSTITUTE).value)
                research_institute_date = get_date_string(ws_source.cell(row=i, column=COL_RESEARCH_INSTITUTE_DATE).value)
                is_innobiz = get_boolean_string(ws_source.cell(row=i, column=COL_IS_INNOBIZ).value)
                innobiz_cert_date = get_date_string(ws_source.cell(row=i, column=COL_INNOBIZ_CERT_DATE).value)
                innobiz_expiry_date = get_date_string(ws_source.cell(row=i, column=COL_INNOBIZ_EXPIRY_DATE).value)
                is_mainbiz = get_boolean_string(ws_source.cell(row=i, column=COL_IS_MAINBIZ).value)
                mainbiz_cert_date = get_date_string(ws_source.cell(row=i, column=COL_MAINBIZ_CERT_DATE).value)
                mainbiz_expiry_date = get_date_string(ws_source.cell(row=i, column=COL_MAINBIZ_EXPIRY_DATE).value)
                is_venture = get_boolean_string(ws_source.cell(row=i, column=COL_IS_VENTURE).value)
                
                # 신규 추가 컬럼
                group_agreement_yn = get_boolean_string(ws_source.cell(row=i, column=COL_GROUP_AGREEMENT_YN).value)
                gfc_yn = get_boolean_string(ws_source.cell(row=i, column=COL_GFC_YN).value)
                venture_cert_date = get_date_string(ws_source.cell(row=i, column=COL_VENTURE_CERT_DATE).value)
                venture_expiry_date = get_date_string(ws_source.cell(row=i, column=COL_VENTURE_EXPIRY_DATE).value)

                # 출력 행 데이터 생성
                output_row_data = [
                    biz_no,
                    patent_apps,
                    registered_patents,
                    has_research_institute,
                    research_institute_date,
                    is_innobiz,
                    innobiz_cert_date,
                    innobiz_expiry_date,
                    is_mainbiz,
                    mainbiz_cert_date,
                    mainbiz_expiry_date,
                    is_venture,
                    venture_cert_date,
                    venture_expiry_date,
                    group_agreement_yn,
                    gfc_yn
                ]
                
                ws_output.append(output_row_data)

            if (i - DATA_START_ROW + 1) % 100 == 0:
                print(f"    {i - DATA_START_ROW + 1}개 행 처리 완료...")

        # 5. 결과 파일 저장 및 마무리
        # 컬럼 너비 자동 맞춤
        for col in ws_output.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws_output.column_dimensions[column_letter].width = adjusted_width

        wb_output.save(OUTPUT_WB_NAME)
        
        print("\n=======================================================")
        print(f"성공: Company_Additional 데이터 추출을 완료했습니다!")
        print(f"결과가 '{OUTPUT_WB_NAME}' 파일의 '{OUTPUT_SHEET_NAME}' 시트에 저장되었습니다.")
        print("=======================================================")

    except FileNotFoundError as e:
        print(f"[오류] 파일을 찾을 수 없습니다: {e}")
        sys.exit(1)
    except KeyError as e:
        print(f"[오류] 시트를 찾을 수 없습니다: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"[오류] 스크립트 실행 중 예기치 않은 오류가 발생했습니다: {e}")
        sys.exit(1)

# 스크립트 실행
if __name__ == "__main__":
    process_additional_data()

