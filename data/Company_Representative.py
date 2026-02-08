import openpyxl
from openpyxl.utils import get_column_letter
import os
import sys
from datetime import datetime

# ====================================================================================
# 파일 및 시트 이름 정의 (사용자 환경에 맞게 이 부분을 수정하세요!)
# ====================================================================================
SOURCE_WB_NAME = "251102_insert.xlsx"
SOURCE_SHEET_NAME = "법인사업자"
OUTPUT_SHEET_NAME = "Company_Representative_Output"
OUTPUT_WB_NAME = "Company_Representative_Output.xlsx" # 결과를 저장할 새 파일 이름

# ====================================================================================
# 엑셀 컬럼 인덱스 정의 (엑셀 파일에 맞게 이 값을 정확히 수정해야 합니다!)
# ====================================================================================
COL_BIZ_NO = 24  # X열

# 대표자 1 정보
COL_REP1_NAME = 43      
COL_REP1_GENDER = 44    
COL_REP1_AGE = 45       
COL_REP1_BIRTH_DATE = 46
COL_REP1_IS_GFC = None  # GFC 필드는 빈값으로 처리

# 대표자 2 정보
COL_REP2_NAME = 51      
COL_REP2_GENDER = 52    
COL_REP2_AGE = 53       
COL_REP2_BIRTH_DATE = 54

# 대표자 3 정보
COL_REP3_NAME = 59      
COL_REP3_GENDER = 60    
COL_REP3_AGE = 61       
COL_REP3_BIRTH_DATE = 62

# --- 상수 정의 ---
DATA_START_ROW = 3

def get_numeric_value(cell_val):
    """
    VBA의 GetNumericValue 함수와 동일한 역할을 합니다.
    셀 값이 숫자 형태가 아니거나 비어있으면 0을 반환합니다.
    """
    if cell_val is None:
        return 0
    try:
        # 정수로 변환 시도
        return int(float(cell_val))
    except (ValueError, TypeError):
        return 0

def convert_age_format(original_age):
    """
    원본 나이를 기반으로 20대~60대 초/중/후반으로 임의 생성합니다.
    예: 원본이 30대면 → 31, 35, 39 중 랜덤하게 선택
    """
    import random
    
    if original_age == 0 or original_age is None:
        # 나이 정보가 없으면 30대~50대 중 랜덤하게 생성
        decade = random.choice([30, 40, 50])
    else:
        # 원본 나이의 십의 자리를 기준으로 연령대 결정
        decade = (original_age // 10) * 10
        
        # 너무 어리거나 나이가 많으면 적절한 범위로 조정
        if decade < 20:
            decade = random.choice([20, 30])
        elif decade > 70:
            decade = random.choice([50, 60, 70])
    
    # 각 연령대별 초/중/후반 중 랜덤 선택
    age_variants = {
        20: [21, 25, 29],  # 20대 초반, 중반, 후반
        30: [31, 35, 39],  # 30대 초반, 중반, 후반
        40: [41, 45, 49],  # 40대 초반, 중반, 후반
        50: [51, 55, 59],  # 50대 초반, 중반, 후반
        60: [61, 65, 69],  # 60대 초반, 중반, 후반
        70: [71, 75, 79]   # 70대 초반, 중반, 후반
    }
    
    # 해당 연령대가 있으면 그 중에서 선택, 없으면 40대 중반으로 기본값
    selected_ages = age_variants.get(decade, [45])
    return random.choice(selected_ages)

def clean_representative_name(name):
    """
    대표자 이름에서 연속된 "*" 문자를 하나로 줄입니다.
    예: "홍**길동" → "홍*길동", "김***철수" → "김*철수"
    """
    if not name:
        return ""
    
    # 연속된 "*" 문자를 하나로 치환
    import re
    cleaned_name = re.sub(r'\*{2,}', '*', name)
    return cleaned_name

def get_date_string(cell_val):
    """
    VBA의 GetDateString 함수와 동일한 역할을 합니다.
    다양한 형태의 날짜를 'YYYYMMDD' 8자리 문자열로 변환합니다.
    """
    if cell_val is None or str(cell_val).strip() == "":
        return ""
    
    # openpyxl이 날짜/시간 객체로 읽은 경우
    if isinstance(cell_val, datetime):
        return cell_val.strftime('%Y%m%d')

    temp_date_str = str(cell_val).strip()
    
    # 6자리 숫자인 경우 (YYMMDD)
    if len(temp_date_str) == 6 and temp_date_str.isdigit():
        yy = int(temp_date_str[:2])
        if 0 <= yy <= 25: # 2000년 ~ 2025년생으로 간주
            return "20" + temp_date_str
        elif 26 <= yy <= 99: # 1926년 ~ 1999년생으로 간주
            return "19" + temp_date_str
    
    # 그 외의 경우는 원본 문자열을 반환
    return temp_date_str


def infer_birth_date(rrn_val, age_group_val):
    """
    대표자 생년월일(46열)과 연령대(45열) 정보를 이용하여 생년월일 6자리를 추론합니다.
    규칙:
      - 46열이 6***** 이고 45열이 60대초반 -> 610101
      - 46열이 6***** 이고 45열이 60대중반 -> 650101
      - 46열이 6***** 이고 45열이 60대후반 -> 690101
      - 4, 5, 6, 7, 8, 9로 시작하는 경우 모두 동일한 규칙 적용
    """
    rrn_str = str(rrn_val).strip() if rrn_val is not None else ""
    age_str = str(age_group_val).strip() if age_group_val is not None else ""
    
    if not rrn_str:
        return ""
    
    # 이미 6자리 이상의 숫자인 경우 (정상 생년월일)
    import re
    if re.match(r'^\d{6}', rrn_str):
        return rrn_str[:6]
    
    # 마스킹된 경우 (예: 6*****) - 첫 글자 추출
    prefix = rrn_str[0]
    if not prefix.isdigit():
        return rrn_str # 숫자가 아니면 그대로 리턴
        
    if prefix not in ['4', '5', '6', '7', '8', '9']:
        return rrn_str # 규칙 적용 대상 아님
    
    # 연령대 텍스트 확인
    suffix = "0101" # 기본값
    if "초반" in age_str:
        suffix = "10101"
    elif "중반" in age_str:
        suffix = "50101"
    elif "후반" in age_str:
        suffix = "90101"
    else:
        # 초/중/후반 정보가 없으면 기본값 (예: X00101 ?)
        # 사용자 요청에 명시되지 않음. 그냥 prefix + 00101? 아니면 그대로?
        # 안전하게 그대로 리턴하거나 기본값 적용. 여기선 10101을 기본으로 함
        suffix = "10101"
        
    return prefix + suffix

def process_representative_data():
    """
    VBA 스크립트(ExtractRepresentativeData)의 핵심 로직을 수행하는 메인 함수입니다.
    """
    try:
        # 1. 엑셀 파일 불러오기
        print(f"'{SOURCE_WB_NAME}' 파일을 불러오는 중입니다...")
        if not os.path.exists(SOURCE_WB_NAME):
            raise FileNotFoundError(f"원본 데이터 파일 '{SOURCE_WB_NAME}'을 찾을 수 없습니다.")

        wb_source = openpyxl.load_workbook(SOURCE_WB_NAME, data_only=True)
        if SOURCE_SHEET_NAME in wb_source.sheetnames:
            ws_source = wb_source[SOURCE_SHEET_NAME]
        else:
            ws_source = wb_source.active
            print(f"Warning: Sheet '{SOURCE_SHEET_NAME}' not found. Using active sheet '{ws_source.title}' instead.")

        # 2. 출력용 새 워크북 및 시트 생성
        wb_output = openpyxl.Workbook()
        ws_output = wb_output.active
        ws_output.title = OUTPUT_SHEET_NAME
        print(f"'{OUTPUT_SHEET_NAME}' 시트를 생성했습니다.")

        # 3. 출력 시트 헤더 작성 (birth_date_inferred 추가?) -> birth_date에 덮어쓰기
        header = ["biz_no", "name", "gender", "age", "birth_date", "is_gfc"]
        ws_output.append(header)
        
        # 4. 데이터 처리 시작
        last_row_source = ws_source.max_row
        print(f"총 {last_row_source - DATA_START_ROW + 1}개의 원본 데이터를 처리합니다...")
        
        extracted_count = 0

        # 대표자별 컬럼 정보를 리스트로 묶어 처리 용이성을 높임
        rep_columns = [
            (COL_REP1_NAME, COL_REP1_GENDER, COL_REP1_AGE, COL_REP1_BIRTH_DATE, COL_REP1_IS_GFC),
            (COL_REP2_NAME, COL_REP2_GENDER, COL_REP2_AGE, COL_REP2_BIRTH_DATE, None), # 대표자2는 GFC 컬럼 없음
            (COL_REP3_NAME, COL_REP3_GENDER, COL_REP3_AGE, COL_REP3_BIRTH_DATE, None)  # 대표자3는 GFC 컬럼 없음
        ]

        for i in range(DATA_START_ROW, last_row_source + 1):
            biz_no_val = ws_source.cell(row=i, column=COL_BIZ_NO).value
            biz_no = str(biz_no_val).strip() if biz_no_val is not None else ""

            if biz_no and biz_no != "0":
                # 대표자 1, 2, 3 정보 순차적으로 처리
                for rep_info in rep_columns:
                    col_name, col_gender, col_age, col_birth, col_gfc = rep_info
                    
                    rep_name_val = ws_source.cell(row=i, column=col_name).value
                    rep_name_raw = str(rep_name_val).strip() if rep_name_val is not None else ""
                    rep_name = clean_representative_name(rep_name_raw)  # 연속된 "*" 문자 정리

                    if rep_name: # 대표자 이름이 있을 경우에만 데이터 추출
                        rep_gender = ws_source.cell(row=i, column=col_gender).value
                        
                        # 나이 계산 (기존 로직 유지)
                        raw_age_val = ws_source.cell(row=i, column=col_age).value
                        raw_age = get_numeric_value(raw_age_val)
                        rep_age = convert_age_format(raw_age)  # 나이 변환 적용
                        
                        # 생년월일 추론 로직 적용
                        rrn_val = ws_source.cell(row=i, column=col_birth).value
                        
                        # raw_age_val (문자열)을 사용하여 추론
                        inferred_birth = infer_birth_date(rrn_val, raw_age_val)
                        
                        if inferred_birth:
                            rep_birth_date = inferred_birth
                        else:
                            # 추론 실패 시 기존 방식(단순 포맷팅) 사용
                            rep_birth_date = get_date_string(rrn_val)
                        
                        # GFC 필드는 항상 빈 공백으로 처리
                        rep_is_gfc = ""

                        output_row_data = [
                            biz_no,
                            rep_name,
                            rep_gender,
                            rep_age,
                            rep_birth_date,
                            rep_is_gfc
                        ]
                        ws_output.append(output_row_data)
                        extracted_count += 1
            
            if (i - DATA_START_ROW + 1) % 100 == 0:
                print(f"    {i - DATA_START_ROW + 1}개 행 처리 완료...")

        # 5. 결과 파일 저장 및 마무리
        # 컬럼 너비 자동 맞춤
        for col in ws_output.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws_output.column_dimensions[column_letter].width = adjusted_width

        wb_output.save(OUTPUT_WB_NAME)
        
        print("\n=======================================================")
        print(f"성공: Company_Representative 데이터 추출을 완료했습니다!")
        print(f"총 {extracted_count}건의 대표자 데이터가 '{OUTPUT_WB_NAME}' 파일에 저장되었습니다.")
        print("=======================================================")

    except FileNotFoundError as e:
        print(f"[오류] 파일을 찾을 수 없습니다: {e}")
        sys.exit(1)
    except KeyError as e:
        print(f"[오류] 시트를 찾을 수 없습니다: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"[오류] 스크립트 실행 중 예기치 않은 오류가 발생했습니다: {e}")
        sys.exit(1)

# 스크립트 실행
if __name__ == "__main__":
    process_representative_data()
