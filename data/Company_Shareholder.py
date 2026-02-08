import openpyxl
from openpyxl.utils import get_column_letter
import os
import sys

# ====================================================================================
# 파일 및 시트 이름 정의 (사용자 환경에 맞게 이 부분을 수정하세요!)
# ====================================================================================
SOURCE_WB_NAME = "251102_insert.xlsx"
SOURCE_SHEET_NAME = "법인사업자"
OUTPUT_SHEET_NAME = "Company_Shareholder_Output"
OUTPUT_WB_NAME = "Company_Shareholder_Output.xlsx"

# --- 상수 정의 ---
DATA_START_ROW = 3
COL_BIZ_NO = 24  # X열

# 대표자 정보 컬럼
COL_REP1_NAME = 43  # AU -> AT (43)
COL_REP2_NAME = 51  # BC -> BB (51)
COL_REP3_NAME = 59  # BK -> BJ (59)

# 주주 정보 시작 컬럼 (각 주주별로 7개 컬럼씩)
# 주주1: 64~70, 주주2: 71~77, 주주3: 78~84, 주주4: 85~91, 주주5: 92~98
COL_SH1_START = 64  # 주주1 시작 컬럼 (이전 65)
COL_SH2_START = 71  # 주주2 시작 컬럼 (이전 72)
COL_SH3_START = 78  # 주주3 시작 컬럼 (이전 79)
COL_SH4_START = 85  # 주주4 시작 컬럼 (이전 86)
COL_SH5_START = 92  # 주주5 시작 컬럼 (이전 93)

# 각 주주 정보 내 컬럼 오프셋
# 주주정보_주주X_이름, 회사와의관계, 실권자와의관계, 소유주식합계, 보통주식수, 우선주식수, 지분율
SHAREHOLDER_NAME_OFFSET = 0          # 이름
COMPANY_RELATIONSHIP_OFFSET = 1      # 회사와의관계
ACTUAL_RELATIONSHIP_OFFSET = 2       # 실권자와의관계
TOTAL_SHARES_OFFSET = 3              # 소유주식합계
COMMON_SHARES_OFFSET = 4             # 보통주식수
PREFERRED_SHARES_OFFSET = 5          # 우선주식수
OWNERSHIP_PERCENT_OFFSET = 6         # 지분율

def get_numeric_value(cell_val):
    """
    셀 값이 숫자 형태가 아니거나 비어있으면 0을 반환합니다.
    """
    if cell_val is None:
        return 0
    try:
        return float(cell_val)
    except (ValueError, TypeError):
        return 0

def clean_shareholder_name(name):
    """
    주주 이름에서 연속된 "*" 문자를 하나로 줄입니다.
    """
    if not name:
        return ""
    
    import re
    cleaned_name = re.sub(r'\*{2,}', '*', name)
    return cleaned_name

def process_shareholder_data():

    """
    기업주주정보 (Company_Shareholder) 추출 메인 함수입니다.
    """
    try:
        # 1. 엑셀 파일 불러오기
        print(f"'{SOURCE_WB_NAME}' 파일을 불러오는 중입니다...")
        if not os.path.exists(SOURCE_WB_NAME):
            raise FileNotFoundError(f"원본 데이터 파일 '{SOURCE_WB_NAME}'을 찾을 수 없습니다.")

        wb_source = openpyxl.load_workbook(SOURCE_WB_NAME, data_only=True)
        if SOURCE_SHEET_NAME in wb_source.sheetnames:
            ws_source = wb_source[SOURCE_SHEET_NAME]
        else:
            ws_source = wb_source.active
            print(f"Warning: Sheet '{SOURCE_SHEET_NAME}' not found. Using active sheet '{ws_source.title}' instead.")

        # 2. 출력용 새 워크북 및 시트 생성
        wb_output = openpyxl.Workbook()
        ws_output = wb_output.active
        ws_output.title = OUTPUT_SHEET_NAME
        print(f"'{OUTPUT_SHEET_NAME}' 시트를 생성했습니다.")

        # 3. 출력 시트 헤더 작성
        headers = ["biz_no", "shareholder_name", "ownership_percent", "relationship", "shareholder_type", "management_type", "silent_partner_relationship"]
        ws_output.append(headers)
        
        # 4. 주주 정보 컬럼 리스트
        shareholder_columns = [
            COL_SH1_START,  # 주주1
            COL_SH2_START,  # 주주2
            COL_SH3_START,  # 주주3
            COL_SH4_START,  # 주주4
            COL_SH5_START   # 주주5
        ]
        
        # 5. 데이터 처리 시작
        last_row_source = ws_source.max_row
        print(f"총 {last_row_source - DATA_START_ROW + 1}개의 원본 데이터를 처리합니다...")

        extracted_count = 0

        for i in range(DATA_START_ROW, last_row_source + 1):
            biz_no_val = ws_source.cell(row=i, column=COL_BIZ_NO).value
            biz_no = str(biz_no_val).strip() if biz_no_val is not None else ""

            if biz_no and biz_no != "0":
                # 대표자 정보 읽기
                rep1_name_val = ws_source.cell(row=i, column=COL_REP1_NAME).value
                rep1_name = str(rep1_name_val).strip() if rep1_name_val is not None else ""
                
                rep2_name_val = ws_source.cell(row=i, column=COL_REP2_NAME).value
                rep2_name = str(rep2_name_val).strip() if rep2_name_val is not None else ""
                
                rep3_name_val = ws_source.cell(row=i, column=COL_REP3_NAME).value
                rep3_name = str(rep3_name_val).strip() if rep3_name_val is not None else ""

                # 각 주주 정보 처리 (주주1~주주5)
                for idx, shareholder_start_col in enumerate(shareholder_columns):
                    # 주주 이름 확인
                    shareholder_name_val = ws_source.cell(row=i, column=shareholder_start_col + SHAREHOLDER_NAME_OFFSET).value
                    shareholder_name_raw = str(shareholder_name_val).strip() if shareholder_name_val is not None else ""
                    shareholder_name = clean_shareholder_name(shareholder_name_raw)

                    if shareholder_name:  # 주주 이름이 있을 경우에만 데이터 추출
                        # 회사와의관계 (컬럼 오프셋 1) [수정]
                        relationship_val = ws_source.cell(row=i, column=shareholder_start_col + COMPANY_RELATIONSHIP_OFFSET).value
                        relationship = str(relationship_val).strip() if relationship_val is not None else ""

                        # 실권자와의관계 (컬럼 오프셋 2)
                        real_relationship_val = ws_source.cell(row=i, column=shareholder_start_col + ACTUAL_RELATIONSHIP_OFFSET).value
                        silent_partner_relationship = str(real_relationship_val).strip() if real_relationship_val is not None else ""
                        
                        # 지분율 (컬럼 오프셋 6)
                        ownership_percent = get_numeric_value(ws_source.cell(row=i, column=shareholder_start_col + OWNERSHIP_PERCENT_OFFSET).value)

                        # 주주 타입 (주주1, 주주2, ...)
                        shareholder_type = f"주주{idx + 1}"
                        
                        # 대표자-주주 동일 여부 확인
                        management_type = ""
                        if shareholder_name == rep1_name or shareholder_name == rep2_name or shareholder_name == rep3_name:
                            management_type = "대표자-주주 동일"

                        output_row_data = [
                            biz_no,
                            shareholder_name,
                            ownership_percent,
                            relationship,
                            shareholder_type,
                            management_type,
                            silent_partner_relationship
                        ]
                        ws_output.append(output_row_data)
                        extracted_count += 1

            if (i - DATA_START_ROW + 1) % 100 == 0:
                print(f"    {i - DATA_START_ROW + 1}개 행 처리 완료...")

        # 6. 결과 파일 저장 및 마무리
        # 컬럼 너비 자동 맞춤
        for col in ws_output.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws_output.column_dimensions[column_letter].width = adjusted_width

        wb_output.save(OUTPUT_WB_NAME)
        
        print("\n=======================================================")
        print(f"성공: Company_Shareholder 데이터 추출을 완료했습니다!")
        print(f"총 {extracted_count}건의 주주 데이터가 '{OUTPUT_WB_NAME}' 파일에 저장되었습니다.")
        print("=======================================================")

    except FileNotFoundError as e:
        print(f"[오류] 파일을 찾을 수 없습니다: {e}")
        sys.exit(1)
    except KeyError as e:
        print(f"[오류] 시트를 찾을 수 없습니다: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"[오류] 스크립트 실행 중 예기치 않은 오류가 발생했습니다: {e}")
        sys.exit(1)

# 스크립트 실행
if __name__ == "__main__":
    process_shareholder_data()



