import openpyxl
from openpyxl.utils import get_column_letter
import os
import sys
from datetime import datetime

# ====================================================================================
# 파일 및 시트 이름 정의 (사용자 환경에 맞게 이 부분을 수정하세요!)
# ====================================================================================
SOURCE_WB_NAME = "251102_insert.xlsx"
SOURCE_SHEET_NAME = "법인사업자"
MAPPING_WB_NAME = "company_basic.xlsx"
MAPPING_SHEET_NAME = "company_basic"
OUTPUT_SHEET_NAME = "Company_Basic_Output"
OUTPUT_WB_NAME = "Company_Basic_Output.xlsx" # 결과를 저장할 새 파일 이름

# --- 상수 정의 ---
DATA_START_ROW = 3
COL_BIZ_NO = 24  # 'X'열

def process_basic_data():
    """
    VBA 스크립트(CreateCompanyBasicCSV)의 핵심 로직을 수행하는 메인 함수입니다.
    """
    try:
        # 1. 엑셀 파일 불러오기
        print(f"'{SOURCE_WB_NAME}' 파일과 '{MAPPING_WB_NAME}' 파일을 불러오는 중입니다...")
        if not os.path.exists(SOURCE_WB_NAME):
            raise FileNotFoundError(f"원본 데이터 파일 '{SOURCE_WB_NAME}'을 찾을 수 없습니다.")
        if not os.path.exists(MAPPING_WB_NAME):
            raise FileNotFoundError(f"매핑 파일 '{MAPPING_WB_NAME}'을 찾을 수 없습니다.")

        wb_source = openpyxl.load_workbook(SOURCE_WB_NAME, data_only=True)
        if SOURCE_SHEET_NAME in wb_source.sheetnames:
            ws_source = wb_source[SOURCE_SHEET_NAME]
        else:
            ws_source = wb_source.active
            print(f"Warning: Sheet '{SOURCE_SHEET_NAME}' not found. Using active sheet '{ws_source.title}' instead.")

        wb_mapping = openpyxl.load_workbook(MAPPING_WB_NAME, data_only=True)
        ws_mapping = wb_mapping[MAPPING_SHEET_NAME]

        # 2. 출력용 새 워크북 및 시트 생성
        wb_output = openpyxl.Workbook()
        ws_output = wb_output.active
        ws_output.title = OUTPUT_SHEET_NAME
        print(f"'{OUTPUT_SHEET_NAME}' 시트를 생성했습니다.")

        # 3. 매핑 정보 읽기 (A열: CSV 헤더, B열: 원본 컬럼 번호)
        print("매핑 정보를 읽는 중입니다...")
        mapping_values = []
        for row in ws_mapping.iter_rows(min_row=2, max_col=2, values_only=True):
            if row[0] is not None and row[1] is not None:
                mapping_values.append(list(row))
        
        # [추가] 필수 컬럼이 매핑 파일에 없는 경우 강제로 추가
        existing_headers = [row[0] for row in mapping_values]
        if "employee_count" not in existing_headers:
             print("Mapping 'employee_count' manually...")
             mapping_values.append(["employee_count", 22]) # 22 is dummy here, logic is in loop
        if "pension_count" not in existing_headers:
             print("Mapping 'pension_count' manually...")
             mapping_values.append(["pension_count", 20]) # 20 is dummy here
        
        # 4. 출력 시트 헤더 작성
        header = [row[0] for row in mapping_values]
        ws_output.append(header)
        
        # 5. 데이터 처리 시작
        last_row_source = ws_source.max_row
        print(f"총 {last_row_source - DATA_START_ROW + 1}개의 원본 데이터를 처리합니다...")

        for i in range(DATA_START_ROW, last_row_source + 1):
            biz_no_val = ws_source.cell(row=i, column=COL_BIZ_NO).value
            biz_no = str(biz_no_val).strip() if biz_no_val is not None else ""

            if biz_no and biz_no != "0":
                output_row_data = []
                for j, mapping_row in enumerate(mapping_values):
                    target_col_name = str(mapping_row[0]).strip()
                    source_col_num = 0
                    
                    # [수정] biz_no 컬럼은 상수 COL_BIZ_NO를 강제로 사용
                    if target_col_name == "biz_no":
                         source_col_num = COL_BIZ_NO  # 24
                    # [추가] 사용자 지정 컬럼 매핑 강제 적용 (master column 기준)
                    elif target_col_name == "company_name":
                         source_col_num = 2  # B열
                    elif target_col_name == "company_size":
                         source_col_num = 3  # C열
                    elif target_col_name == "company_type":
                         source_col_num = 4  # D열
                    elif target_col_name == "corporate_reg_no":
                         source_col_num = 25  # Y열
                    elif target_col_name == "establish_date":
                         source_col_num = 26  # Z열
                    elif target_col_name == "industry_code":
                         source_col_num = 31  # AE열
                    elif target_col_name == "industry_name":
                         source_col_num = 32  # AF열
                    elif target_col_name == "zip_code":
                         source_col_num = 33  # AG열
                    elif target_col_name == "address":
                         source_col_num = 34  # AH열
                    elif target_col_name == "region":
                         source_col_num = 35  # AI열
                    elif target_col_name == "city_district":
                         source_col_num = 36  # AJ열
                    elif target_col_name == "phone_number":
                         source_col_num = 37  # AK열
                    elif target_col_name == "email":
                         source_col_num = 38  # AL열
                    elif target_col_name == "fax_number":
                         source_col_num = 39  # AM열
                    elif target_col_name == "representative_name":
                         source_col_num = 43  # AQ열
                    elif target_col_name == "national_pension":
                         source_col_num = 20  # T열
                    elif target_col_name == "pension_count" or target_col_name == "employee_count":
                         source_col_num = 22  # V열 (근로자수)
                    elif target_col_name == "pension_date":
                         source_col_num = 23  # W열
                    elif target_col_name == "ked_transaction_yn":
                         source_col_num = 160  # 벤처인증 거래 유무
                    elif target_col_name == "patent_transaction_yn":
                         source_col_num = 163  # 특허 거래 유무
                    elif target_col_name == "group_transaction_yn":
                         source_col_num = 166  # 단체 거래 유무
                    elif target_col_name == "gfc_transaction_yn":
                         source_col_num = 167  # GFC 거래 유무
                    else:
                        try:
                            # 매핑 파일의 컬럼 번호가 숫자인지 확인
                            source_col_num = int(mapping_row[1])
                        except (ValueError, TypeError):
                            # 숫자가 아니면 해당 셀은 비워둠
                            output_row_data.append("")
                            continue

                    cell_value = ws_source.cell(row=i, column=source_col_num).value
                    
                    # 날짜 형식 특별 처리
                    if target_col_name in ["establish_date", "employee_count_date"]:
                        # openpyxl은 날짜를 datetime 객체로 읽어옵니다.
                        if isinstance(cell_value, datetime):
                            processed_value = cell_value.strftime('%Y-%m-%d')
                        else:
                            processed_value = cell_value # 날짜 형식이 아니면 그대로 사용
                    else:
                        processed_value = cell_value

                    output_row_data.append(processed_value)
                
                ws_output.append(output_row_data)

            if (i - DATA_START_ROW + 1) % 100 == 0:
                print(f"    {i - DATA_START_ROW + 1}개 행 처리 완료...")

        # 6. 결과 파일 저장 및 마무리
        # 컬럼 너비 자동 맞춤
        for col in ws_output.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws_output.column_dimensions[column_letter].width = adjusted_width

        wb_output.save(OUTPUT_WB_NAME)
        
        # 7. CSV 파일도 생성
        import pandas as pd
        
        # Excel 파일을 다시 읽어서 CSV로 저장
        df = pd.read_excel(OUTPUT_WB_NAME, sheet_name=OUTPUT_SHEET_NAME)
        csv_filename = OUTPUT_WB_NAME.replace('.xlsx', '.csv')
        df.to_csv(csv_filename, index=False, encoding='utf-8')
        
        print("\n=======================================================")
        print(f"성공: Company_Basic 데이터 추출을 완료했습니다!")
        print(f"결과가 '{OUTPUT_WB_NAME}' 파일의 '{OUTPUT_SHEET_NAME}' 시트에 저장되었습니다.")
        print(f"CSV 파일도 '{csv_filename}'에 저장되었습니다.")
        print("=======================================================")

    except FileNotFoundError as e:
        print(f"[오류] 파일을 찾을 수 없습니다: {e}")
        sys.exit(1)
    except KeyError as e:
        print(f"[오류] 시트를 찾을 수 없습니다: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"[오류] 스크립트 실행 중 예기치 않은 오류가 발생했습니다: {e}")
        sys.exit(1)

# 스크립트 실행
if __name__ == "__main__":
    process_basic_data()
