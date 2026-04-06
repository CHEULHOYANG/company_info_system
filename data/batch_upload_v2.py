# -*- coding: utf-8 -*-
"""
Integrated Batch Data Upload Script (v2) - Enhanced
- Supports XLSX input for data tables
- Modes: 'validate' (generation + dry-run/count) and 'execute' (DB update)
- Returns JSON statistics for backend parsing
"""

import os
import sys
import subprocess
import time
import sqlite3
import pandas as pd
import shutil
import json
import argparse
from datetime import datetime

# Configuration
DATA_DIR = os.path.dirname(os.path.abspath(__file__))
# DB_PATH relative to data directory (one level up)
DB_PATH = os.path.join(os.path.dirname(DATA_DIR), "company_database.db")

# Scripts to run for Generation
GENERATION_SCRIPTS = [
    "basic.py",                    # Company_Basic
    "fanance.py",                  # Company_Financial
    "Company_Additional.py",       # Company_Additional
    "Company_Representative.py",   # Company_Representative
    "Company_Shareholder.py"       # Company_Shareholder
]

# File Mapping (Preferred: XLSX, Fallback: CSV)
# Key matches the table name text used in logic
TABLE_FILES = {
    "Company_Basic": ["Company_Basic_Output.xlsx", "Company_Basic_Output.csv"],
    "Company_Financial": ["Company_Financial_Output.xlsx", "Company_Financial_Output.csv"],
    "Company_Shareholder": ["Company_Shareholder_Output.xlsx", "Company_Shareholder_Output.csv"],
    "Company_Representative": ["Company_Representative_Output.xlsx", "Company_Representative_Output.csv"],
    "Company_Additional": ["Company_Additional_Output.xlsx", "Company_Additional_Output.csv"]
}

def log(message):
    timestamp = datetime.now().strftime("%H:%M:%S")
    # Prefix regular logs so frontend can filter if needed, 
    # but for now we just print.
    print(f"[{timestamp}] {message}", flush=True)

def json_result(key, value):
    """Print a JSON object on a separate line for the backend to parse"""
    print(json.dumps({"type": "result", "key": key, "value": value}), flush=True)

def progress_event(event_type, table, **kwargs):
    """Print a progress event for real-time UI updates"""
    data = {"type": event_type, "table": table}
    data.update(kwargs)
    print(json.dumps(data), flush=True)

def cleanup_old_files():
    """Delete existing Output files to ensure fresh generation"""
    log("Cleaning up old Output files...")
    for key, files in TABLE_FILES.items():
        for filename in files:
            if os.path.exists(filename):
                try:
                    os.remove(filename)
                    # log(f"  Deleted old {filename}")
                except Exception as e:
                    log(f"  Warning: Could not delete {filename}: {e}")

def check_source_file():
    source_file = "251102_insert.xlsx"
    if os.path.exists(source_file):
        size_mb = os.path.getsize(source_file) / (1024 * 1024)
        mtime = datetime.fromtimestamp(os.path.getmtime(source_file)).strftime('%Y-%m-%d %H:%M:%S')
        log(f"Source File: {source_file} ({size_mb:.2f} MB)")
        return True
    else:
        log(f"ERROR: Source file {source_file} not found!")
        return False

def run_generation_scripts():
    """Optimized: Executes the conversion logic using Pandas directly for speed"""
    log("=== Step 1: generating Output files from Excel (Optimized) ===")
    source_file = "251102_insert.xlsx"
    
    try:
        from openpyxl import load_workbook
        
        log(f"Loading workbook (read-only mode): {source_file}")
        wb = load_workbook(source_file, read_only=True, data_only=True)
        
        # --- Auto-Detect Sheet and Columns ---
        sheet_name = "법인사업자" if "법인사업자" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]
        log(f"Source Sheet: {sheet_name}")

        rows_gen = ws.rows
        detected_biz_col = 24
        found_header_row = 3 # 1-indexed

        # Scan first 20 rows for "사업자" header
        header_rows = []
        for r_idx, row in enumerate(ws.iter_rows(max_row=20), 1):
            found_in_row = False
            for c_idx, cell in enumerate(row, 1):
                val = str(cell.value or "").strip()
                if "사업자" in val and ("번호" in val or "등록" in val):
                    detected_biz_col = c_idx
                    found_header_row = r_idx
                    log(f"Auto-detected BIZ_NO column: {detected_biz_col} at Row: {found_header_row} ('{val}')")
                    found_in_row = True
                    break
            if found_in_row: break

        # Data starts after header row
        def get_row_data(row_obj):
            return [cell.value for cell in row_obj]

        def get_v(row_data, idx):
            try:
                if idx <= 0 or idx > len(row_data): return None
                v = row_data[idx-1]
                return v if v is not None and str(v).strip().lower() != "none" else None
            except: return None

        def get_n(v, m=1000):
            try:
                if v is None: return 0
                s = str(v).replace(',', '').strip()
                if not s or s == "None": return 0
                return float(s) * m
            except: return 0

        # Manual Mapping
        mapping_dict = {
            "biz_no": detected_biz_col, "company_name": 2, "company_size": 3, "company_type": 4,
            "corporate_reg_no": 25, "establish_date": 26, "industry_code": 31,
            "industry_name": 32, "zip_code": 33, "jibun_address": 34, "address": 35,
            "region": 36, "city_district": 37, "phone_number": 38, "email": 39,
            "fax_number": 40, "representative_name": 44, "national_pension": 20,
            "pension_count": 22, "employee_count": 22, "pension_date": 23,
            "ked_transaction_yn": 161, "patent_transaction_yn": 164, "group_transaction_yn": 167,
            "gfc_transaction_yn": 168
        }

        basic_rows = []; fin_rows = []; rep_rows = []; sha_rows = []; add_rows = []
        
        # Process all rows in one pass
        count = 0
        log("Processing rows...")
        
        for r_idx, row in enumerate(ws.iter_rows(min_row=found_header_row + 1)):
            row_data = get_row_data(row)
            biz_no = get_v(row_data, detected_biz_col)
            if not biz_no: continue
            biz_no = str(biz_no).strip()
            count += 1
            
            if count % 100 == 0:
                log(f"Progress: {count} records extracted...")

            # 1. Company_Basic
            basic_item = []
            basic_headers = list(mapping_dict.keys())
            for h in basic_headers:
                col_idx = mapping_dict[h]
                val = get_v(row_data, col_idx)
                if h == "establish_date" and isinstance(val, (datetime, pd.Timestamp)):
                    val = val.strftime('%Y-%m-%d')
                basic_item.append(val)
            basic_rows.append(basic_item)

            # 2. Company_Financial
            r1 = str(get_v(row_data, 7) or ""); r2 = str(get_v(row_data, 6) or ""); r3 = str(get_v(row_data, 5) or "")
            ta = get_n(get_v(row_data, 139)); tl = get_n(get_v(row_data, 137)); cs = get_n(get_v(row_data, 138))
            er = get_n(get_v(row_data, 125)); apic = get_n(get_v(row_data, 126)); la = get_n(get_v(row_data, 117))
            ba = get_n(get_v(row_data, 118)); irg = get_n(get_v(row_data, 119)); irb = get_n(get_v(row_data, 120))
            ri = get_n(get_v(row_data, 121)); ramt = get_n(get_v(row_data, 122)); adp = get_n(get_v(row_data, 127))
            adr = get_n(get_v(row_data, 128)); csv = get_n(get_v(row_data, 129)); ure = get_n(get_v(row_data, 130))
            ca = get_n(get_v(row_data, 131)); ce = get_n(get_v(row_data, 132)); stl = get_n(get_v(row_data, 133))
            std = get_n(get_v(row_data, 134)); sltb = get_n(get_v(row_data, 135)); ii = get_n(get_v(row_data, 136))
            cr = get_n(get_v(row_data, 138)); sic = get_n(get_v(row_data, 140), 1)

            years = [(2024, 100, 101, 102, 123, 144), (2023, 103, 104, 105, 123, 145), (2022, 106, 107, 108, 123, 146)]
            for yr, s_col, o_col, e_col, r_col, t_col in years:
                sales = get_n(get_v(row_data, s_col), 1000000); op_inc = get_n(get_v(row_data, o_col), 1000000)
                te = get_n(get_v(row_data, e_col)); ret = get_n(get_v(row_data, r_col)); tax = get_n(get_v(row_data, t_col))
                fin_rows.append([biz_no, yr, r1, r2, r3, sales, op_inc, op_inc, ta, tl, te, ret, cs, er, apic, tax, la, ba, irg, irb, ri, ramt, adp, adr, csv, ure, ca, ce, stl, std, sltb, ii, cr, sic])

            # 3. Company_Representative
            name = str(get_v(row_data, 44) or "").strip()
            if name:
                gender = "M" if "남" in str(get_v(row_data, 45) or "") else ("F" if "여" in str(get_v(row_data, 45) or "") else "")
                age = int(get_n(get_v(row_data, 46), 1))
                rep_rows.append([biz_no, name, gender, age, "N", ""])

            # 4. Company_Shareholder
            sh_reps = [str(get_v(row_data, 44) or "").strip(), str(get_v(row_data, 52) or "").strip(), str(get_v(row_data, 60) or "").strip()]
            for sh_idx, s_col in enumerate([65, 72, 79, 86, 93]):
                sh_nm = str(get_v(row_data, s_col) or "").strip()
                if sh_nm and sh_nm != "0":
                    sh_rel = str(get_v(row_data, s_col+1) or "").strip()
                    sh_sil = str(get_v(row_data, s_col+2) or "").strip()
                    sh_p = get_n(get_v(row_data, s_col+6), 1)
                    m_type = "대표자-주주 동일" if sh_nm in sh_reps else ""
                    sha_rows.append([biz_no, sh_nm, sh_p, sh_rel, f"주주{sh_idx+1}", m_type, sh_sil])

            # 5. Company_Additional
            def get_b(v): 
                s = str(v or "").strip().upper()
                return "1" if s in ["Y", "1", "YES", "TRUE", "여"] else "0"
            def get_d(v):
                if isinstance(v, (datetime, pd.Timestamp)): return v.strftime('%Y-%m-%d')
                s_val = str(v or "").strip()
                return s_val if s_val != "0" else ""
            
            add_rows.append([
                biz_no, get_n(get_v(row_data, 166), 1), get_n(get_v(row_data, 165), 1),
                get_b(get_v(row_data, 153)), get_d(get_v(row_data, 154)),
                get_b(get_v(row_data, 155)), get_d(get_v(row_data, 156)), get_d(get_v(row_data, 157)),
                get_b(get_v(row_data, 158)), get_d(get_v(row_data, 159)), get_d(get_v(row_data, 160)),
                get_b(get_v(row_data, 161)), get_d(get_v(row_data, 162)), get_d(get_v(row_data, 163)),
                get_b(get_v(row_data, 167)), get_b(get_v(row_data, 168))
            ])

        log(f"Extraction complete: {count} total valid records.")
        
        # Save Outputs using Pandas (writing is faster than mapping)
        log("Saving validation files...")
        if basic_rows: pd.DataFrame(basic_rows, columns=list(mapping_dict.keys())).to_excel("Company_Basic_Output.xlsx", index=False)
        if fin_rows: pd.DataFrame(fin_rows, columns=["biz_no", "fiscal_year", "rating1", "rating2", "rating3", "sales_revenue", "operating_income", "net_income", "total_assets", "total_liabilities", "total_equity", "retained_earnings", "capital_surplus", "earned_reserve", "additional_paid_in_capital", "corporate_tax", "land_asset", "building_asset", "investment_real_ground", "investment_real_building", "rental_income", "rent_amt", "advances_paid", "advances_received", "capital_stock_value", "undistributed_retained_earnings", "current_assets", "cash_equivalents", "short_term_loan", "short_term_deposit", "principal_short_long_term_bonds", "interest_income", "capital_reserve", "shares_issued_count"]).to_excel("Company_Financial_Output.xlsx", index=False)
        if rep_rows: pd.DataFrame(rep_rows, columns=["biz_no", "name", "gender", "age", "is_gfc", "birth_date"]).to_excel("Company_Representative_Output.xlsx", index=False)
        if sha_rows: pd.DataFrame(sha_rows, columns=["biz_no", "shareholder_name", "ownership_percent", "relationship", "shareholder_type", "management_type", "silent_partner_relationship"]).to_excel("Company_Shareholder_Output.xlsx", index=False)
        if add_rows: pd.DataFrame(add_rows, columns=["biz_no", "patent_applications_count", "registered_patents_count", "has_research_institute", "research_institute_date", "is_innobiz", "innobiz_cert_date", "innobiz_expiry_date", "is_mainbiz", "mainbiz_cert_date", "mainbiz_expiry_date", "is_venture", "venture_cert_date", "venture_expiry_date", "group_agreement_yn", "gfc_yn"]).to_excel("Company_Additional_Output.xlsx", index=False)
        log(f"Validation complete: Found {count} records across all categories.")
        return True

    except Exception as e:
        log(f"CRITICAL ERROR in optimized generation: {e}")
        import traceback
        log(traceback.format_exc())
        return False

def load_data_frame(table_name):
    """Smart loader: tries XLSX then CSV"""
    files = TABLE_FILES.get(table_name, [])
    for f in files:
        if os.path.exists(f):
            try:
                if f.endswith('.xlsx'):
                    return pd.read_excel(f)
                else:
                    return pd.read_csv(f) # Default utf-8
            except Exception as e:
                log(f"Error reading {f}: {e}")
    return None

def ensure_base_tables(conn):
    """Create necessary tables if they don't exist"""
    cursor = conn.cursor()
    
    # Company_Basic
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS Company_Basic (
            biz_no VARCHAR(12) PRIMARY KEY,
            company_name VARCHAR(255),
            company_size VARCHAR(50),
            company_type VARCHAR(50),
            industry_code VARCHAR(10),
            industry_name VARCHAR(255),
            zip_code VARCHAR(10),
            jibun_address VARCHAR(500),
            address VARCHAR(500),
            region VARCHAR(50),
            city_district VARCHAR(50),
            phone_number VARCHAR(20),
            email VARCHAR(255),
            fax_number VARCHAR(20),
            gfc_transaction_yn CHAR(1),
            group_transaction_yn CHAR(1),
            ked_transaction_yn CHAR(1),
            patent_transaction_yn CHAR(1),
            establish_date DATE,
            corporate_reg_no VARCHAR(20),
            representative_name VARCHAR(100),
            national_pension INTEGER,
            pension_count INTEGER,
            pension_date DATE
        )
    ''')

    # Company_Financial
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS Company_Financial (
            biz_no VARCHAR(12),
            fiscal_year INTEGER,
            rating1 VARCHAR(20),
            rating2 VARCHAR(20),
            rating3 VARCHAR(20),
            sales_revenue BIGINT,
            operating_income BIGINT,
            net_income BIGINT,
            total_assets BIGINT,
            total_liabilities BIGINT,
            total_equity BIGINT,
            retained_earnings BIGINT,
            capital_surplus BIGINT,
            earned_reserve BIGINT,
            additional_paid_in_capital BIGINT,
            corporate_tax BIGINT,
            land_asset BIGINT,
            building_asset BIGINT,
            investment_real_ground BIGINT,
            investment_real_building BIGINT,
            rental_income BIGINT,
            rent_amt BIGINT,
            advances_paid BIGINT,
            advances_received BIGINT,
            capital_stock_value BIGINT,
            undistributed_retained_earnings BIGINT,
            current_assets BIGINT,
            cash_equivalents BIGINT,
            short_term_loan BIGINT,
            short_term_deposit BIGINT,
            principal_short_long_term_bonds BIGINT,
            interest_income BIGINT,
            capital_reserve BIGINT,
            shares_issued_count INTEGER,
            PRIMARY KEY (biz_no, fiscal_year),
            FOREIGN KEY (biz_no) REFERENCES Company_Basic (biz_no)
        )
    ''')

    # Company_Shareholder
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS Company_Shareholder (
            company_shareholder_id INTEGER PRIMARY KEY,
            biz_no VARCHAR(12),
            shareholder_name VARCHAR(100),
            ownership_percent NUMERIC,
            relationship VARCHAR(100),
            shareholder_type VARCHAR(50),
            management_type VARCHAR(50),
            silent_partner_relationship VARCHAR(100),
            total_shares_owned BIGINT,
            FOREIGN KEY (biz_no) REFERENCES Company_Basic (biz_no)
        )
    ''')

    # Company_Representative
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS Company_Representative (
            company_representative_id INTEGER PRIMARY KEY,
            biz_no VARCHAR(12),
            name VARCHAR(100),
            gender CHAR(1),
            age INTEGER,
            birth_date DATE,
            is_gfc CHAR(1),
            FOREIGN KEY (biz_no) REFERENCES Company_Basic (biz_no)
        )
    ''')

    # Company_Additional
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS Company_Additional (
            company_additional_id INTEGER PRIMARY KEY,
            biz_no VARCHAR(12),
            patent_applications_count INTEGER,
            registered_patents_count INTEGER,
            has_research_institute CHAR(1),
            research_institute_date DATE,
            is_innobiz CHAR(1),
            innobiz_cert_date DATE,
            innobiz_expiry_date DATE,
            is_mainbiz CHAR(1),
            mainbiz_cert_date DATE,
            mainbiz_expiry_date DATE,
            is_venture CHAR(1),
            venture_cert_date DATE,
            venture_expiry_date DATE,
            FOREIGN KEY (biz_no) REFERENCES Company_Basic (biz_no)
        )
    ''')
    conn.commit()

    # [Migration] Ensure jibun_address exists in Company_Basic
    cursor.execute("PRAGMA table_info(Company_Basic)")
    cols = [row[1] for row in cursor.fetchall()]
    if "jibun_address" not in cols:
        try:
            cursor.execute("ALTER TABLE Company_Basic ADD COLUMN jibun_address VARCHAR(500)")
            conn.commit()
            log("Added jibun_address column to Company_Basic table via migration.")
        except Exception as e:
            log(f"Migration error: {e}")

def connect_db():
    try:
        conn = sqlite3.connect(DB_PATH, timeout=60)
        conn.execute("PRAGMA journal_mode=WAL")
        ensure_base_tables(conn)
        return conn
    except Exception as e:
        log(f"Database connection failed: {e}")
        return None

def is_masked(name):
    return '*' in str(name)

def fuzzy_name_match(masked_name, real_name):
    if not masked_name or not real_name:
        return False
    s_masked = str(masked_name).strip()
    s_real = str(real_name).strip()
    if len(s_masked) < 2 or len(s_real) < 2:
        return s_masked == s_real
    return s_masked[0] == s_real[0] and s_masked[-1] == s_real[-1]

# ---------------- Processing Functions ----------------

def process_company_basic(conn, df, execute=False, global_current=0, global_total=0):
    log("Processing Company_Basic...")
    if 'biz_no' in df.columns:
        df['biz_no'] = df['biz_no'].astype(str).str.replace(r'[\s-]', '', regex=True)
    
    total_rows = len(df)
    log(f"  Found {total_rows} rows")
    json_result("Company_Basic_Found", total_rows)
    
    if conn is None: return global_current

    # Emit table start event
    progress_event("table_start", "Company_Basic", total=total_rows)

    cursor = conn.cursor()
    count_updated = 0
    count_inserted = 0
    count_error = 0
    
    cursor.execute("PRAGMA table_info(Company_Basic)")
    db_cols = [row[1] for row in cursor.fetchall()]

    for i, (_, row) in enumerate(df.iterrows()):
        global_current += 1
        # Emit progress every 50 rows globally
        if global_current % 50 == 0 or i == 0 or i == total_rows - 1:
            progress_event("progress", "Company_Basic", current=global_current, total=global_total)
            
        biz_no = row.get('biz_no')
        if not biz_no:
            count_error += 1
            if count_error == 1: json_result("LastError", f"[Company_Basic] BizNo missing. Row: {row.to_dict()}")
            continue
        
        try:
            cursor.execute("SELECT 1 FROM Company_Basic WHERE biz_no = ?", (biz_no,))
            exists = cursor.fetchone()
            
            data_to_save = {}
            for col in db_cols:
                if col in row:
                    data_to_save[col] = row[col]
            
            if exists:
                # UPDATE
                count_updated += 1
                if execute:
                    update_data = {k: v for k, v in data_to_save.items() if k != 'biz_no'}
                    if update_data:
                        set_clause = ", ".join([f"{k} = ?" for k in update_data.keys()])
                        values = list(update_data.values()) + [biz_no]
                        cursor.execute(f"UPDATE Company_Basic SET {set_clause} WHERE biz_no = ?", values)
            else:
                # INSERT
                count_inserted += 1
                if execute:
                    cols = list(data_to_save.keys())
                    placeholders = ", ".join(["?" for _ in cols])
                    values = list(data_to_save.values())
                    cursor.execute(f"INSERT INTO Company_Basic ({', '.join(cols)}) VALUES ({placeholders})", values)
        except Exception as e:
            count_error += 1
            if count_error == 1: json_result("LastError", f"[Company_Basic] {str(e)}")
            log(f"  [Error] Basic biz_no={biz_no}: {e}")
            
    if execute: conn.commit()
    log(f"  Inserted: {count_inserted}, Updated: {count_updated}, Errors: {count_error}")
    json_result("Company_Basic_Inserted", count_inserted)
    json_result("Company_Basic_Updated", count_updated)
    json_result("Company_Basic_Error", count_error)
    progress_event("table_done", "Company_Basic", inserted=count_inserted, updated=count_updated, error=count_error)
    return global_current

def process_company_representative(conn, df, execute=False, global_current=0, global_total=0):
    log("Processing Company_Representative...")
    if 'biz_no' in df.columns:
        df['biz_no'] = df['biz_no'].astype(str).str.replace(r'[\s-]', '', regex=True)
    
    total_rows = len(df)
    log(f"  Found {total_rows} rows")
    json_result("Company_Representative_Found", total_rows)
    
    if conn is None: return global_current

    cursor = conn.cursor()
    count_inserted = 0
    count_updated = 0
    count_error = 0
    
    # Emit table start event
    progress_event("table_start", "Company_Representative", total=total_rows)
    
    for i, (_, row) in enumerate(df.iterrows()):
        global_current += 1
        if global_current % 50 == 0 or i == 0 or i == total_rows - 1:
            progress_event("progress", "Company_Representative", current=global_current, total=global_total)

        biz_no = row.get('biz_no')
        new_name = str(row.get('name', '')).strip()
        if not new_name:
            count_error += 1
            if count_error == 1: json_result("LastError", f"[Company_Representative] Name missing. Row: {row.to_dict()}")
            continue
            
        try:
            cursor.execute("SELECT company_representative_id, name FROM Company_Representative WHERE biz_no = ?", (biz_no,))
            existing_reps = cursor.fetchall()
            
            match_found = False
            target_id = None
            
            # 1. Exact Match
            for rid, rname in existing_reps:
                if rname == new_name:
                    match_found = True; target_id = rid; break
            
            # 2. Masked Match
            if not match_found and is_masked(new_name):
                for rid, rname in existing_reps:
                    if not is_masked(rname) and fuzzy_name_match(new_name, rname):
                        match_found = True; target_id = rid
                        row['name'] = rname # Keep DB name
                        break
            
            if target_id:
                # Update (simple fields)
                count_updated += 1
                if execute:
                    sql = "UPDATE Company_Representative SET gender=?, age=?, is_gfc=? WHERE company_representative_id=?"
                    cursor.execute(sql, (row.get('gender',''), row.get('age',0), row.get('is_gfc','N'), target_id))
            else:
                # Insert
                count_inserted += 1
                if execute:
                    sql = "INSERT INTO Company_Representative (biz_no, name, gender, age, is_gfc, birth_date) VALUES (?, ?, ?, ?, ?, ?)"
                    cursor.execute(sql, (biz_no, new_name, row.get('gender',''), row.get('age',0), row.get('is_gfc','N'), row.get('birth_date','')))
        except Exception as e:
            count_error += 1
            if count_error == 1: json_result("LastError", f"[Company_Representative] {str(e)}")
            log(f"  [Error] Rep biz_no={biz_no}: {e}")
        
    if execute: conn.commit()
    log(f"  Inserted: {count_inserted}, Updated: {count_updated}, Errors: {count_error}")
    json_result("Company_Representative_Inserted", count_inserted)
    json_result("Company_Representative_Updated", count_updated)
    json_result("Company_Representative_Error", count_error)
    progress_event("table_done", "Company_Representative", inserted=count_inserted, updated=count_updated, error=count_error)
    return global_current

def process_company_shareholder(conn, df, execute=False, global_current=0, global_total=0):
    log("Processing Company_Shareholder...")
    if 'biz_no' in df.columns:
        df['biz_no'] = df['biz_no'].astype(str).str.replace(r'[\s-]', '', regex=True)
        
    total_rows = len(df)
    log(f"  Found {total_rows} rows")
    json_result("Company_Shareholder_Found", total_rows)
    
    if conn is None: return global_current

    cursor = conn.cursor()
    count_inserted = 0
    cols = [row[1] for row in cursor.fetchall()]
    if "jibun_address" not in cols:
        try:
            cursor.execute("ALTER TABLE Company_Basic ADD COLUMN jibun_address VARCHAR(500)")
            conn.commit()
            log("Added jibun_address column to Company_Basic table via migration.")
        except Exception as e:
            log(f"Migration error: {e}")

def connect_db():
    try:
        conn = sqlite3.connect(DB_PATH, timeout=60)
        conn.execute("PRAGMA journal_mode=WAL")
        ensure_base_tables(conn)
        return conn
    except Exception as e:
        log(f"Database connection failed: {e}")
        return None

def is_masked(name):
    return '*' in str(name)

def fuzzy_name_match(masked_name, real_name):
    if not masked_name or not real_name:
        return False
    s_masked = str(masked_name).strip()
    s_real = str(real_name).strip()
    if len(s_masked) < 2 or len(s_real) < 2:
        return s_masked == s_real
    return s_masked[0] == s_real[0] and s_masked[-1] == s_real[-1]

# ---------------- Processing Functions ----------------

def process_company_basic(conn, df, execute=False, global_current=0, global_total=0):
    log("Processing Company_Basic...")
    if 'biz_no' in df.columns:
        df['biz_no'] = df['biz_no'].astype(str).str.replace(r'[\s-]', '', regex=True)
    
    total_rows = len(df)
    log(f"  Found {total_rows} rows")
    json_result("Company_Basic_Found", total_rows)
    
    if conn is None: return global_current

    # Emit table start event
    progress_event("table_start", "Company_Basic", total=total_rows)

    cursor = conn.cursor()
    count_updated = 0
    count_inserted = 0
    count_error = 0
    
    cursor.execute("PRAGMA table_info(Company_Basic)")
    db_cols = [row[1] for row in cursor.fetchall()]

    for i, (_, row) in enumerate(df.iterrows()):
        global_current += 1
        # Emit progress every 50 rows globally
        if global_current % 50 == 0 or i == 0 or i == total_rows - 1:
            progress_event("progress", "Company_Basic", current=global_current, total=global_total)
        
        # Log every 500 rows for visibility in console
        if (i + 1) % 500 == 0:
            log(f"  > [Company_Basic] Processing {i+1} / {total_rows}...")
            
        biz_no = row.get('biz_no')
        if not biz_no:
            count_error += 1
            if count_error == 1: json_result("LastError", f"[Company_Basic] BizNo missing. Row: {row.to_dict()}")
            continue
        
        try:
            cursor.execute("SELECT 1 FROM Company_Basic WHERE biz_no = ?", (biz_no,))
            exists = cursor.fetchone()
            
            data_to_save = {}
            for col in db_cols:
                if col in row:
                    data_to_save[col] = row[col]
            
            if exists:
                # UPDATE
                count_updated += 1
                if execute:
                    update_data = {k: v for k, v in data_to_save.items() if k != 'biz_no'}
                    if update_data:
                        set_clause = ", ".join([f"{k} = ?" for k in update_data.keys()])
                        values = list(update_data.values()) + [biz_no]
                        cursor.execute(f"UPDATE Company_Basic SET {set_clause} WHERE biz_no = ?", values)
            else:
                # INSERT
                count_inserted += 1
                if execute:
                    cols = list(data_to_save.keys())
                    placeholders = ", ".join(["?" for _ in cols])
                    values = list(data_to_save.values())
                    cursor.execute(f"INSERT INTO Company_Basic ({', '.join(cols)}) VALUES ({placeholders})", values)
        except Exception as e:
            count_error += 1
            if count_error == 1: json_result("LastError", f"[Company_Basic] {str(e)}")
            log(f"  [Error] Basic biz_no={biz_no}: {e}")
            
    if execute: conn.commit()
    log(f"  Inserted: {count_inserted}, Updated: {count_updated}, Errors: {count_error}")
    json_result("Company_Basic_Inserted", count_inserted)
    json_result("Company_Basic_Updated", count_updated)
    json_result("Company_Basic_Error", count_error)
    progress_event("table_done", "Company_Basic", inserted=count_inserted, updated=count_updated, error=count_error)
    return global_current

def process_company_representative(conn, df, execute=False, global_current=0, global_total=0):
    log("Processing Company_Representative...")
    if 'biz_no' in df.columns:
        df['biz_no'] = df['biz_no'].astype(str).str.replace(r'[\s-]', '', regex=True)
    
    total_rows = len(df)
    log(f"  Found {total_rows} rows")
    json_result("Company_Representative_Found", total_rows)
    
    if conn is None: return global_current

    cursor = conn.cursor()
    count_inserted = 0
    count_updated = 0
    count_error = 0
    
    # Emit table start event
    progress_event("table_start", "Company_Representative", total=total_rows)
    
    for i, (_, row) in enumerate(df.iterrows()):
        global_current += 1
        if global_current % 50 == 0 or i == 0 or i == total_rows - 1:
            progress_event("progress", "Company_Representative", current=global_current, total=global_total)
            
        if (i + 1) % 500 == 0:
            log(f"  > [Company_Representative] Processing {i+1} / {total_rows}...")

        biz_no = row.get('biz_no')
        new_name = str(row.get('name', '')).strip()
        if not new_name:
            count_error += 1
            if count_error == 1: json_result("LastError", f"[Company_Representative] Name missing. Row: {row.to_dict()}")
            continue
            
        try:
            cursor.execute("SELECT company_representative_id, name FROM Company_Representative WHERE biz_no = ?", (biz_no,))
            existing_reps = cursor.fetchall()
            
            match_found = False
            target_id = None
            
            # 1. Exact Match
            for rid, rname in existing_reps:
                if rname == new_name:
                    match_found = True; target_id = rid; break
            
            # 2. Masked Match
            if not match_found and is_masked(new_name):
                for rid, rname in existing_reps:
                    if not is_masked(rname) and fuzzy_name_match(new_name, rname):
                        match_found = True; target_id = rid
                        row['name'] = rname # Keep DB name
                        break
            
            if target_id:
                # Update (simple fields)
                count_updated += 1
                if execute:
                    sql = "UPDATE Company_Representative SET gender=?, age=?, is_gfc=? WHERE company_representative_id=?"
                    cursor.execute(sql, (row.get('gender',''), row.get('age',0), row.get('is_gfc','N'), target_id))
            else:
                # Insert
                count_inserted += 1
                if execute:
                    sql = "INSERT INTO Company_Representative (biz_no, name, gender, age, is_gfc, birth_date) VALUES (?, ?, ?, ?, ?, ?)"
                    cursor.execute(sql, (biz_no, new_name, row.get('gender',''), row.get('age',0), row.get('is_gfc','N'), row.get('birth_date','')))
        except Exception as e:
            count_error += 1
            if count_error == 1: json_result("LastError", f"[Company_Representative] {str(e)}")
            log(f"  [Error] Rep biz_no={biz_no}: {e}")
        
    if execute: conn.commit()
    log(f"  Inserted: {count_inserted}, Updated: {count_updated}, Errors: {count_error}")
    json_result("Company_Representative_Inserted", count_inserted)
    json_result("Company_Representative_Updated", count_updated)
    json_result("Company_Representative_Error", count_error)
    progress_event("table_done", "Company_Representative", inserted=count_inserted, updated=count_updated, error=count_error)
    return global_current

def process_company_shareholder(conn, df, execute=False, global_current=0, global_total=0):
    log("Processing Company_Shareholder...")
    if 'biz_no' in df.columns:
        df['biz_no'] = df['biz_no'].astype(str).str.replace(r'[\s-]', '', regex=True)
        
    total_rows = len(df)
    log(f"  Found {total_rows} rows")
    json_result("Company_Shareholder_Found", total_rows)
    
    if conn is None: return global_current

    cursor = conn.cursor()
    count_inserted = 0
    count_updated = 0
    count_error = 0
    
    # Emit table start event
    progress_event("table_start", "Company_Shareholder", total=total_rows)
     # 1. Identify companies in this batch
    unique_biz_nos = df['biz_no'].unique().tolist()
    if not unique_biz_nos:
        return global_current
        
    if execute:
        log(f"  Deleting existing shareholders for {len(unique_biz_nos)} companies...")
        DELETE_CHUNK = 500
        for i in range(0, len(unique_biz_nos), DELETE_CHUNK):
            chunk = unique_biz_nos[i:i + DELETE_CHUNK]
            placeholders = ','.join(['?'] * len(chunk))
            cursor.execute(f"DELETE FROM Company_Shareholder WHERE biz_no IN ({placeholders})", chunk)
    
    # 2. Bulk Insert New Data
    log("  Processing shareholder data entries...")
    
    CHUNK_SIZE = 100
    insert_data_list = []
    insert_cols = ['biz_no', 'shareholder_name', 'ownership_percent', 'relationship', 
                   'shareholder_type', 'management_type', 'silent_partner_relationship', 'total_shares_owned']
                   
    for i, (_, row) in enumerate(df.iterrows()):
        global_current += 1
        if global_current % 50 == 0 or i == 0 or i == total_rows - 1:
            progress_event("progress", "Company_Shareholder", current=global_current, total=global_total)
        
        if (i+1) % 500 == 0:
            log(f"  > [Company_Shareholder] Processing {i+1} / {total_rows}...")
            
        biz_no = row.get('biz_no')
        csv_name = str(row.get('shareholder_name', '')).strip()
        if not csv_name:
            count_error += 1
            continue

        csv_stock = 0
        if 'total_shares_owned' in row:
             try: csv_stock = int(float(str(row['total_shares_owned']).replace(',', '')))
             except: csv_stock = 0

        vals = [biz_no, csv_name, row.get('ownership_percent', 0), row.get('relationship', ''),
                row.get('shareholder_type', ''), row.get('management_type', ''), row.get('silent_partner_relationship', ''), csv_stock]
        
        insert_data_list.append(vals)
        count_inserted += 1

        if execute and len(insert_data_list) >= CHUNK_SIZE:
             placeholders = ", ".join(["?" for _ in insert_cols])
             cursor.executemany(f"INSERT INTO Company_Shareholder ({', '.join(insert_cols)}) VALUES ({placeholders})", insert_data_list)
             insert_data_list = []
             conn.commit()

    if execute and insert_data_list:
        try:
            placeholders = ", ".join(["?" for _ in insert_cols])
            cursor.executemany(f"INSERT INTO Company_Shareholder ({', '.join(insert_cols)}) VALUES ({placeholders})", insert_data_list)
        except Exception as e:
            log(f"  [Error] Final Bulk Insert failed: {e}")
            count_error += len(insert_data_list)

    if execute: conn.commit()
    log(f"  Processed: {count_inserted}, Errors: {count_error}")
    json_result("Company_Shareholder_Inserted", count_inserted)
    json_result("Company_Shareholder_Updated", 0)
    json_result("Company_Shareholder_Error", count_error)
    progress_event("table_done", "Company_Shareholder", inserted=count_inserted, updated=0, error=count_error)
    return global_current

def process_company_financial(conn, df, execute=False, global_current=0, global_total=0):
    log("Processing Company_Financial...")
    if 'biz_no' in df.columns:
        df['biz_no'] = df['biz_no'].astype(str).str.replace(r'[\s-]', '', regex=True)
    
    total_rows = len(df)
    log(f"  Found {total_rows} rows")
    json_result("Company_Financial_Found", total_rows)
    
    if conn is None: return global_current

    cursor = conn.cursor()
    cursor.execute("PRAGMA table_info(Company_Financial)")
    db_cols = [row[1] for row in cursor.fetchall() if row[1] != 'company_financial_id']

    count_inserted = 0
    count_updated = 0
    count_error = 0
    
    keys_to_delete = df[['biz_no', 'fiscal_year']].drop_duplicates().values.tolist()
    if execute and keys_to_delete:
        log(f"  Deleting existing financials for {len(keys_to_delete)} records...")
        DELETE_CHUNK = 500
        for i in range(0, len(keys_to_delete), DELETE_CHUNK):
            chunk = keys_to_delete[i:i + DELETE_CHUNK]
            cursor.executemany("DELETE FROM Company_Financial WHERE biz_no = ? AND fiscal_year = ?", chunk)
        
    log("  Processing financial data entries...")
    CHUNK_SIZE = 100
    insert_data_list = []
    insert_cols = [col for col in db_cols if col != 'company_financial_id']
    
    for i, (_, row) in enumerate(df.iterrows()):
        global_current += 1
        if global_current % 50 == 0 or i == 0 or i == total_rows - 1:
            progress_event("progress", "Company_Financial", current=global_current, total=global_total)
        
        if (i+1) % 500 == 0:
            log(f"  > [Company_Financial] Processing {i+1} / {total_rows}...")
        
        biz_no = row.get('biz_no')
        fiscal_year = row.get('fiscal_year')
        if not biz_no or not fiscal_year:
            count_error += 1
            continue

        data_list = []
        for col in insert_cols:
            val = row.get(col)
            if pd.isna(val): val = None
            elif hasattr(val, 'item'): val = val.item()
            data_list.append(val)
            
        insert_data_list.append(data_list)
        count_inserted += 1

        if execute and len(insert_data_list) >= CHUNK_SIZE:
             placeholders = ", ".join(["?" for _ in insert_cols])
             cursor.executemany(f"INSERT INTO Company_Financial ({', '.join(insert_cols)}) VALUES ({placeholders})", insert_data_list)
             insert_data_list = []
             conn.commit()

    if execute and insert_data_list:
        try:
            placeholders = ", ".join(["?" for _ in insert_cols])
            cursor.executemany(f"INSERT INTO Company_Financial ({', '.join(insert_cols)}) VALUES ({placeholders})", insert_data_list)
        except Exception as e:
            log(f"  [Error] Final Bulk Insert failed: {e}")
            count_error += len(insert_data_list)

    if execute: conn.commit()
    log(f"  Processed: {count_inserted}, Errors: {count_error}")
    json_result("Company_Financial_Inserted", count_inserted)
    json_result("Company_Financial_Updated", 0)
    json_result("Company_Financial_Error", count_error)
    progress_event("table_done", "Company_Financial", inserted=count_inserted, updated=0, error=count_error)
    return global_current

def process_generic_table(conn, df, table_name, pk_cols, execute=False, global_current=0, global_total=0):
    log(f"Processing {table_name}...")
    if 'biz_no' in df.columns:
        df['biz_no'] = df['biz_no'].astype(str).str.replace(r'[\s-]', '', regex=True)
    
    total_rows = len(df)
    log(f"  Found {total_rows} rows")
    json_result(f"{table_name}_Found", total_rows)
    
    if conn is None: return global_current

    cursor = conn.cursor()
    cursor.execute(f"PRAGMA table_info({table_name})")
    db_cols = [row[1] for row in cursor.fetchall()]

    count_inserted = 0
    count_updated = 0
    count_error = 0
    
    for i, (_, row) in enumerate(df.iterrows()):
        global_current += 1
        if global_current % 50 == 0 or i == 0 or i == total_rows - 1:
            progress_event("progress", table_name, current=global_current, total=global_total)
        
        if (i+1) % 500 == 0:
            log(f"  > [{table_name}] Processing {i+1} / {total_rows}...")
        
        try:
            where_parts = [f"{pk} = ?" for pk in pk_cols]; where_vals = [row[pk] for pk in pk_cols]
            cursor.execute(f"SELECT count(*) FROM {table_name} WHERE {' AND '.join(where_parts)}", where_vals)
            exists = cursor.fetchone()[0] > 0
            
            data = {col: row[col] for col in db_cols if col in row}
            if exists:
                count_updated += 1
                if execute:
                    update_data = {k: v for k, v in data.items() if k not in pk_cols}
                    if update_data:
                        set_parts = [f"{k}=?" for k in update_data.keys()]
                        sql = f"UPDATE {table_name} SET {', '.join(set_parts)} WHERE {' AND '.join(where_parts)}"
                        cursor.execute(sql, list(update_data.values()) + where_vals)
            else:
                count_inserted += 1
                if execute:
                    cols = list(data.keys()); placeholders = ",".join(["?" for _ in cols])
                    sql = f"INSERT INTO {table_name} ({','.join(cols)}) VALUES ({placeholders})"
                    cursor.execute(sql, list(data.values()))
        except Exception as e:
            count_error += 1
            log(f"  [Error] {table_name} row {i}: {e}")
        
    if execute: conn.commit()
    log(f"  Inserted: {count_inserted}, Updated: {count_updated}, Errors: {count_error}")
    json_result(f"{table_name}_Inserted", count_inserted)
    json_result(f"{table_name}_Updated", count_updated)
    json_result(f"{table_name}_Error", count_error)
    progress_event("table_done", table_name, inserted=count_inserted, updated=count_updated, error=count_error)
    return global_current

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--mode", choices=['validate', 'execute'], default='validate')
    parser.add_argument("--skip-gen", action="store_true")
    parser.add_argument("--db-path", type=str)
    args = parser.parse_args()

    mode = args.mode
    global DB_PATH
    if args.db_path: DB_PATH = args.db_path
    
    log(f"=== Batch Upload Started (Mode: {mode}) ===")
    if not check_source_file(): return
    if not args.skip_gen:
        cleanup_old_files()
        if not run_generation_scripts(): return

    log(f"=== Step 2: Processing Data (Mode: {mode}) ===")
    conn = connect_db()
    if not conn: return

    table_data = {
        "Company_Basic": load_data_frame("Company_Basic"),
        "Company_Representative": load_data_frame("Company_Representative"),
        "Company_Shareholder": load_data_frame("Company_Shareholder"),
        "Company_Financial": load_data_frame("Company_Financial"),
        "Company_Additional": load_data_frame("Company_Additional")
    }

    global_total = sum(len(df) for df in table_data.values() if df is not None)
    global_current = 0
    log(f"Global Update: Total {global_total} rows.")

    if table_data["Company_Basic"] is not None:
        global_current = process_company_basic(conn, table_data["Company_Basic"], execute=(mode == 'execute'), global_current=global_current, global_total=global_total)
    if table_data["Company_Representative"] is not None:
        global_current = process_company_representative(conn, table_data["Company_Representative"], execute=(mode == 'execute'), global_current=global_current, global_total=global_total)
    if table_data["Company_Shareholder"] is not None:
        global_current = process_company_shareholder(conn, table_data["Company_Shareholder"], execute=(mode == 'execute'), global_current=global_current, global_total=global_total)
    if table_data["Company_Financial"] is not None:
        global_current = process_company_financial(conn, table_data["Company_Financial"], execute=(mode == 'execute'), global_current=global_current, global_total=global_total)
    if table_data["Company_Additional"] is not None:
        global_current = process_generic_table(conn, table_data["Company_Additional"], "Company_Additional", ["biz_no"], execute=(mode == 'execute'), global_current=global_current, global_total=global_total)

    if conn:
        conn.close()
    
    # Ensure 100% progress at the very end
    progress_event("progress", "Complete", current=global_total, total=global_total)
    log("=== Batch Upload Complete ===")

if __name__ == "__main__":
    main()
