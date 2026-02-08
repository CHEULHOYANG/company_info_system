import openpyxl
from openpyxl.utils import get_column_letter
import os
import sys

# ====================================================================================
# Configuration
# ====================================================================================
SOURCE_WB_NAME = "251102_insert.xlsx"
SOURCE_SHEET_NAME = "법인사업자"
OUTPUT_SHEET_NAME = "Company_Financial_Output"
OUTPUT_WB_NAME = "Company_Financial_Output.xlsx"

# Constants
DATA_START_ROW = 3
COL_BIZ_NO = 24  # 'X' column

# Fiscal Years (Assumed based on column order 99->102->105)
# 99: Current (2024), 102: Previous (2023), 105: Before Previous (2022)
YEARS = [2024, 2023, 2022]

def get_numeric_value(cell_val):
    if cell_val is None:
        return 0
    try:
        return float(cell_val)
    except (ValueError, TypeError):
        return 0

def get_string_value(cell_val):
    if cell_val is None:
        return ""
    return str(cell_val).strip()

def process_financial_data():
    try:
        print(f"Loading '{SOURCE_WB_NAME}'...")
        if not os.path.exists(SOURCE_WB_NAME):
            raise FileNotFoundError(f"Source file '{SOURCE_WB_NAME}' not found.")

        wb_source = openpyxl.load_workbook(SOURCE_WB_NAME, data_only=True)
        if SOURCE_SHEET_NAME in wb_source.sheetnames:
            ws_source = wb_source[SOURCE_SHEET_NAME]
        else:
            ws_source = wb_source.active
            print(f"Warning: Sheet '{SOURCE_SHEET_NAME}' not found. Using active sheet '{ws_source.title}' instead.")

        # Create Output Workbook
        wb_output = openpyxl.Workbook()
        ws_output = wb_output.active
        ws_output.title = OUTPUT_SHEET_NAME
        print(f"Created sheet '{OUTPUT_SHEET_NAME}'.")

        # Headers matching the new schema
        headers = [
            "biz_no",
            "fiscal_year",
            "rating1", "rating2", "rating3",
            "sales_revenue", "operating_income", "net_income",
            "total_assets", "total_liabilities", "total_equity",
            "retained_earnings", "capital_surplus", "earned_reserve",
            "additional_paid_in_capital", "corporate_tax",
            "land_asset", "building_asset",
            "investment_real_ground", "investment_real_building",
            "rental_income", "rent_amt",
            "advances_paid", "advances_received",
            "capital_stock_value", "undistributed_retained_earnings",
            "current_assets", "cash_equivalents",
            "short_term_loan", "short_term_deposit",
            "principal_short_long_term_bonds", "interest_income",
            "capital_reserve", "shares_issued_count"
        ]
        ws_output.append(headers)

        last_row_source = ws_source.max_row
        print(f"Processing {last_row_source - DATA_START_ROW + 1} rows from source...")

        for i in range(DATA_START_ROW, last_row_source + 1):
            biz_no_val = ws_source.cell(row=i, column=COL_BIZ_NO).value
            biz_no = str(biz_no_val).strip() if biz_no_val is not None else ""

            if biz_no and biz_no != "0":
                # Extract Shared Data (Columns that are same for all years or single values)
                # Ratings: 7, 6, 5
                rating1 = get_string_value(ws_source.cell(row=i, column=7).value)
                rating2 = get_string_value(ws_source.cell(row=i, column=6).value)
                rating3 = get_string_value(ws_source.cell(row=i, column=5).value)
                
                # Single Values (Assuming same for all 3 years based on request structure)
                # e.g. "138번째 * 1000, 138번째 * 1000, 138번째 * 1000"
                total_assets = get_numeric_value(ws_source.cell(row=i, column=138).value) * 1000
                total_liabilities = get_numeric_value(ws_source.cell(row=i, column=136).value) * 1000
                capital_surplus = get_numeric_value(ws_source.cell(row=i, column=137).value) * 1000 # 137,137,137
                earned_reserve = get_numeric_value(ws_source.cell(row=i, column=124).value) * 1000 # 124,124,124
                additional_paid_in_capital = get_numeric_value(ws_source.cell(row=i, column=125).value) * 1000 # 125,125,125
                
                land_asset = get_numeric_value(ws_source.cell(row=i, column=116).value) * 1000
                building_asset = get_numeric_value(ws_source.cell(row=i, column=117).value) * 1000
                investment_real_ground = get_numeric_value(ws_source.cell(row=i, column=118).value) * 1000
                investment_real_building = get_numeric_value(ws_source.cell(row=i, column=119).value) * 1000
                rental_income = get_numeric_value(ws_source.cell(row=i, column=120).value) * 1000
                rent_amt = get_numeric_value(ws_source.cell(row=i, column=121).value) * 1000
                
                advances_paid = get_numeric_value(ws_source.cell(row=i, column=126).value) * 1000
                advances_received = get_numeric_value(ws_source.cell(row=i, column=127).value) * 1000
                capital_stock_value = get_numeric_value(ws_source.cell(row=i, column=128).value) * 1000
                undistributed_retained_earnings = get_numeric_value(ws_source.cell(row=i, column=129).value) * 1000
                
                current_assets = get_numeric_value(ws_source.cell(row=i, column=130).value) * 1000
                cash_equivalents = get_numeric_value(ws_source.cell(row=i, column=131).value) * 1000
                short_term_loan = get_numeric_value(ws_source.cell(row=i, column=132).value) * 1000
                short_term_deposit = get_numeric_value(ws_source.cell(row=i, column=133).value) * 1000
                principal_short_long_term_bonds = get_numeric_value(ws_source.cell(row=i, column=134).value) * 1000
                interest_income = get_numeric_value(ws_source.cell(row=i, column=135).value) * 1000
                capital_reserve = get_numeric_value(ws_source.cell(row=i, column=137).value) * 1000 # Same as capital_surplus? User said 137 again.
                shares_issued_count = get_numeric_value(ws_source.cell(row=i, column=139).value)

                # Process 3 Years
                # Year 1 (2024): 99, 100, 101, 122, 143
                # Year 2 (2023): 102, 103, 104, 122, 144
                # Year 3 (2022): 105, 106, 107, 122, 145
                
                year_cols = [
                    # Year 1
                    {
                        'year': YEARS[0],
                        'sales': 99, 'op_inc': 100, 'net_inc': 100, # Wait, user said "net_income: 100,103,106". Operating also 100,103,106??
                        # User Request:
                        # sales_revenue: 99, 102, 105
                        # operating_income: 100, 103, 106
                        # net_income: 100, 103, 106 -> Duplicate? Usually Net Income is different.
                        # Let's check headers. If 100 is Op Income, Net might be 101?
                        # But User explicitly said "100,103,106" for both.
                        # Wait, "operating_income ... 100,103,106"
                        # "net_income ... 100,103,106"
                        # This looks like a copy-paste error in user request OR they are the same in this file?
                        # "total_equity ... 101, 104, 107"
                        # "retained_earnings ... 122, 122, 122" (Shared?)
                        # "corporate_tax ... 143, 144, 145"
                        
                        # I will follow user request strictly. If they map to same column, so be it.
                        'equity': 101, 'retained': 122, 'corp_tax': 143
                    },
                    # Year 2
                    {
                        'year': YEARS[1],
                        'sales': 102, 'op_inc': 103, 'net_inc': 103,
                        'equity': 104, 'retained': 122, 'corp_tax': 144
                    },
                    # Year 3
                    {
                        'year': YEARS[2],
                        'sales': 105, 'op_inc': 106, 'net_inc': 106,
                        'equity': 107, 'retained': 122, 'corp_tax': 145
                    }
                ]

                for y_data in year_cols:
                    fiscal_year = y_data['year']
                    
                    sales_revenue = get_numeric_value(ws_source.cell(row=i, column=y_data['sales']).value) * 1000000
                    operating_income = get_numeric_value(ws_source.cell(row=i, column=y_data['op_inc']).value) * 1000000
                    net_income = get_numeric_value(ws_source.cell(row=i, column=y_data['net_inc']).value) * 1000000
                    total_equity = get_numeric_value(ws_source.cell(row=i, column=y_data['equity']).value) * 1000
                    retained_earnings = get_numeric_value(ws_source.cell(row=i, column=y_data['retained']).value) * 1000
                    corporate_tax = get_numeric_value(ws_source.cell(row=i, column=y_data['corp_tax']).value) * 1000
                    
                    output_row = [
                        biz_no,
                        fiscal_year,
                        rating1, rating2, rating3,
                        sales_revenue, operating_income, net_income,
                        total_assets, total_liabilities, total_equity,
                        retained_earnings, capital_surplus, earned_reserve,
                        additional_paid_in_capital, corporate_tax,
                        land_asset, building_asset,
                        investment_real_ground, investment_real_building,
                        rental_income, rent_amt,
                        advances_paid, advances_received,
                        capital_stock_value, undistributed_retained_earnings,
                        current_assets, cash_equivalents,
                        short_term_loan, short_term_deposit,
                        principal_short_long_term_bonds, interest_income,
                        capital_reserve, shares_issued_count
                    ]
                    ws_output.append(output_row)

            if (i - DATA_START_ROW + 1) % 100 == 0:
                print(f"    {i - DATA_START_ROW + 1} rows processed...")

        for col in ws_output.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws_output.column_dimensions[column].width = adjusted_width

        wb_output.save(OUTPUT_WB_NAME)
        print(f"\nSuccess! Saved to '{OUTPUT_WB_NAME}'.")

    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    process_financial_data()
