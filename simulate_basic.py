import openpyxl
import os

SOURCE_WB_NAME = r"F:\ai_data\GFCAISearch_경기안양시_최신근로자수10인이상_제조_260203_1118.xlsx"
DATA_START_ROW = 3
COL_BIZ_NO = 24

try:
    print(f"Loading {SOURCE_WB_NAME}...")
    wb = openpyxl.load_workbook(SOURCE_WB_NAME, data_only=True)
    ws = wb.active
    print(f"Active sheet: {ws.title}")
    
    print(f"Reading from row {DATA_START_ROW}...")
    for i in range(DATA_START_ROW, DATA_START_ROW + 5):
        val = ws.cell(row=i, column=COL_BIZ_NO).value
        print(f"Row {i}, Col {COL_BIZ_NO} (BizNo): {val}")
        
    # Search for specific biz no mentioned by user
    target_biz = "1078150926"
    print(f"Searching for {target_biz} in first 100 rows...")
    found = False
    for i in range(1, 100):
        # Check all columns just in case
        for j in range(1, 50):
            val = ws.cell(row=i, column=j).value
            if val and str(val).replace('-','').strip() == target_biz:
                print(f"FOUND {target_biz} at Row {i}, Col {j}")
                found = True
    
    if not found:
        print(f"Did not find {target_biz} in first 100 rows.")

except Exception as e:
    print(f"Error: {e}")
